"""엑셀 파일 저장 모듈."""

import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config


def save_to_excel(trades: list[dict], file_path: str, title: str, date_str: str):
    """거래 데이터를 엑셀 파일로 저장.

    Args:
        trades: 거래 데이터 리스트
        file_path: 엑셀 파일 경로
        title: 시트 타이틀
        date_str: 날짜 문자열 (시트명)
    """
    os.makedirs(config.DATA_DIR, exist_ok=True)

    if not trades:
        print("[Excel] 저장할 데이터가 없습니다.")
        return

    # DataFrame 생성
    df = pd.DataFrame(trades)

    # 컬럼 순서 정리
    ordered_cols = [c for c in config.COLUMN_NAMES if c in df.columns]
    df = df[ordered_cols]

    # 컬럼명을 표시명으로 변경
    df = df.rename(columns=config.COLUMN_NAMES)

    # 시트명
    sheet_name = date_str

    # 기존 파일이 있으면 시트 추가/교체
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            del book[sheet_name]
            print(f"[Excel] 기존 시트 '{sheet_name}' 삭제 후 재생성")
        book.save(file_path)

        with pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    else:
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    _apply_styles(file_path, sheet_name, title, df)

    print(f"[Excel] 저장 완료: {file_path} (시트: {sheet_name})")


def _format_value(value):
    """금액을 읽기 쉬운 형식으로 변환."""
    if not isinstance(value, (int, float)) or pd.isna(value):
        return ""
    value = float(value)
    if abs(value) >= 1e9:
        return f"${value / 1e9:.2f}B"
    if abs(value) >= 1e6:
        return f"${value / 1e6:.2f}M"
    if abs(value) >= 1e3:
        return f"${value / 1e3:.1f}K"
    return f"${value:,.2f}"


def _apply_styles(file_path: str, sheet_name: str, title: str, df: pd.DataFrame):
    """엑셀 시트에 스타일 적용."""
    book = load_workbook(file_path)
    ws = book[sheet_name]

    num_cols = len(df.columns)

    # 타이틀 행 (1행)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{title} - {sheet_name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 헤더 스타일 (2행)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Total Value 컬럼을 읽기 쉬운 형식으로 변환
    if "Total Value" in df.columns:
        val_col = list(df.columns).index("Total Value") + 1
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=val_col)
            cell.value = _format_value(cell.value)

    # 열 너비 자동 조정
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # 데이터 영역 숫자 정렬
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=num_cols):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")

    # 자동 필터
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}{ws.max_row}"

    # 틀 고정 (헤더 아래)
    ws.freeze_panes = "A3"

    book.save(file_path)
